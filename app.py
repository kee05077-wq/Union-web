from __future__ import annotations

import atexit
import os
import subprocess
import sys
import time
import re
from pathlib import Path

# TensorFlow의 oneDNN 안내 로그 등 불필요한 C++ 레벨 로그 억제.
# mediapipe/tensorflow가 import되기 전에 설정해야 적용된다.
os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "3")
os.environ.setdefault("GLOG_minloglevel", "3")

PROJECT_DIR = Path(__file__).resolve().parent
REQUIREMENTS_PATH = PROJECT_DIR / "requirements.txt"

# tensorflow==2.21.0 / mediapipe==0.10.35 는 Python 3.11 전용 휠만 검증됨
REQUIRED_PYTHON = (3, 11)
if sys.version_info[:2] != REQUIRED_PYTHON:
    print(f"[ERROR] 이 프로젝트는 Python {REQUIRED_PYTHON[0]}.{REQUIRED_PYTHON[1]} 전용입니다. (현재: Python {sys.version_info[0]}.{sys.version_info[1]})")
    print(f"        tensorflow/mediapipe 고정 버전이 지원하지 않는 Python 버전입니다. 'py -{REQUIRED_PYTHON[0]}.{REQUIRED_PYTHON[1]} app.py' 로 실행해 주세요.")
    input("Press Enter to close...")
    raise SystemExit(1)

def install_missing_packages(missing_package: str) -> None:
    print(f"[SYSTEM] Missing package detected: {missing_package}")
    subprocess.run([sys.executable, "-m", "ensurepip", "--upgrade"], cwd=PROJECT_DIR, check=False)
    subprocess.run([sys.executable, "-m", "pip", "install", "-r", str(REQUIREMENTS_PATH)], cwd=PROJECT_DIR, check=True)

try:
    from flask import Flask, jsonify, request, send_from_directory, render_template
    from flask_socketio import SocketIO, emit, join_room, leave_room
    from openpyxl import Workbook, load_workbook
    from collections import deque
    from engineio.payload import Payload
    import base64
    import threading
    import eventlet
    import eventlet.tpool
    import cv2
    import numpy as np
    from tensorflow.keras.models import load_model
    from unicode import join_jamos, CHAR_MEDIALS
    from modules.hand_module import HandDetector
    from modules.utils import Vector_Normalization, resample_time_buffer
except ModuleNotFoundError as error:
    missing_package = getattr(error, "name", "required package")
    try:
        install_missing_packages(missing_package)
        subprocess.run([sys.executable, str(PROJECT_DIR / "app.py")], cwd=PROJECT_DIR, check=False)
        raise SystemExit(0)
    except Exception as install_error:
        print(f"[ERROR] Python package not found: {missing_package}")
        input("Press Enter to close...")
        raise SystemExit(1)

# ==========================================
# 기본 설정
# ==========================================
DATABASE_DIR = PROJECT_DIR / "database"
WORKBOOK_PATH = DATABASE_DIR / "users.xlsx"
SHEET_NAME = "Users"
HTML_FILES = {"index.html", "chatentry.html", "room.html", "story.html", "signup.html", "recover.html", "translate.html", "emergency.html"}

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.config['SECRET_KEY'] = 'union-secret-key'
socketio = SocketIO(app, cors_allowed_origins="*", max_http_buffer_size=1e7, async_mode='eventlet')

# sign_frame이 200ms 간격으로 계속 들어오는 동안 MediaPipe/TensorFlow 처리가 지연되면
# 폴링 트랜스포트에 패킷이 쌓여 기본 한도(16개)를 넘기므로 여유 있게 올려둔다.
Payload.max_decode_packets = 200

# 수어 큐 및 타임아웃 관리
user_finger_queues = {}    # 확정되어 큐에 편입된 자모
user_pending_jamo = {}     # 아직 확정되지 않은 후보 음운 (자음<->모음 전환 전까지 갱신만 됨)
user_sentence_words = {}   # 확정되어 문장에 편입된 단어 목록
user_last_active = {}

CHAR_MEDIALS_SET = set(CHAR_MEDIALS)

def is_vowel(jamo: str) -> bool:
    return jamo in CHAR_MEDIALS_SET

WORD_PAUSE_SECONDS = 1.5      # 이 시간만큼 손 인식이 끊기면 진행 중이던 자모를 하나의 단어로 확정
SENTENCE_PAUSE_SECONDS = 1.5  # 단어 확정과 동시에 즉시 문장으로 묶어 전송 (1.5초 단일 임계값)

# 정규식 필터 (단일 자모 제거)
def remove_single_letters(word: str) -> str:
    return re.sub(r'[ㄱ-ㅎㅏ-ㅣ]', '', word)

# ==========================================
# 손 제스처 분류 모델 로드 (MediaPipe Holistic + Keras)
# ==========================================
try:
    print("[SYSTEM] 수어 제스처 분류 모델(gesture_classifier.h5)을 로드합니다...")
    model = load_model(PROJECT_DIR / "gesture_classifier.h5", compile=False)
except Exception as e:
    print(f"[ERROR] 모델 로드 실패: {e}")

# Keras 모델은 여러 스레드에서 동시에 predict()를 호출하는 상황을 보장하지 않으므로 직렬화한다.
predict_lock = threading.Lock()

ACTIONS = ['ㄱ', 'ㄴ', 'ㄷ', 'ㄹ', 'ㅁ', 'ㅂ', 'ㅅ', 'ㅇ', 'ㅈ', 'ㅊ', 'ㅋ', 'ㅌ', 'ㅍ', 'ㅎ',
           'ㅏ', 'ㅑ', 'ㅓ', 'ㅕ', 'ㅗ', 'ㅛ', 'ㅜ', 'ㅠ', 'ㅡ', 'ㅣ',
           'ㅐ', 'ㅒ', 'ㅔ', 'ㅖ', 'ㅢ', 'ㅚ', 'ㅟ']
SEQ_LENGTH = 10
PREDICTION_CONFIDENCE_THRESHOLD = 0.5

# 학습 데이터(record_dataset.py)는 웹캠의 실제 fps(보통 30)로 녹화되었으므로,
# SEQ_LENGTH 프레임은 학습 시점 기준 약 WINDOW_SECONDS 만큼의 실제 시간에 해당한다.
# 실시간 소켓 프레임은 도착 간격이 불규칙하므로, 프레임 "개수"가 아니라 이 시간 창을
# 기준으로 리샘플링해야 학습 때와 같은 시간축을 모델에 넣을 수 있다.
TRAIN_FPS = 30.0
WINDOW_SECONDS = SEQ_LENGTH / TRAIN_FPS

# 세션(room+name)별 프레임 시퀀스 버퍼와 MediaPipe 검출기
# 각 원소는 (timestamp, feature_vector) 튜플이며 시간 오름차순으로 쌓인다.
user_landmark_seqs: dict[str, deque] = {}
user_detectors: dict[str, HandDetector] = {}

def get_detector(queue_key: str) -> HandDetector:
    if queue_key not in user_detectors:
        user_detectors[queue_key] = HandDetector(min_detection_confidence=0.5)
    return user_detectors[queue_key]

def _detect_right_hand(detector: HandDetector, img):
    img = detector.findHands(img)
    return detector.findRightHandLandmark(img)

def _predict_gesture(input_data):
    with predict_lock:
        return model.predict(input_data, verbose=0)[0]

def reset_sessions() -> None:
    """서버 시작/종료 시 세션별 자모 조합 상태와 MediaPipe 검출기를 깨끗하게 초기화한다."""
    for detector in user_detectors.values():
        try:
            detector.landmarker.close()
        except Exception:
            pass

    user_finger_queues.clear()
    user_pending_jamo.clear()
    user_sentence_words.clear()
    user_last_active.clear()
    user_landmark_seqs.clear()
    user_detectors.clear()

# ==========================================
# 엑셀 DB 함수
# ==========================================
def ensure_workbook() -> None:
    DATABASE_DIR.mkdir(parents=True, exist_ok=True)
    if WORKBOOK_PATH.exists(): return
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(["name", "birthDate", "username", "password", "createdAt"])
    workbook.save(WORKBOOK_PATH)

def read_users() -> list[dict[str, str]]:
    ensure_workbook()
    workbook = load_workbook(WORKBOOK_PATH)
    worksheet = workbook[SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows: return []
    headers = [str(value or "").strip() for value in rows[0]]
    users: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(row): continue
        users.append({headers[index]: str(value or "").strip() for index, value in enumerate(row) if index < len(headers)})
    return users

def write_users(users: list[dict[str, str]]) -> None:
    ensure_workbook()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    headers = ["name", "birthDate", "username", "password", "createdAt"]
    worksheet.append(headers)
    for user in users:
        worksheet.append([user.get(header, "") for header in headers])
    workbook.save(WORKBOOK_PATH)
    workbook.close()

def same_text(left: str | None, right: str | None) -> bool:
    return str(left or "").strip() == str(right or "").strip()

# ==========================================
# 라우터 및 API
# ==========================================
@app.get("/")
def serve_index(): return render_template("index.html")

@app.get("/<path:filename>")
def serve_root_file(filename: str):
    if filename in HTML_FILES: return render_template(filename)
    return send_from_directory(PROJECT_DIR, filename)

@app.post("/api/signup")
def signup():
    payload = request.get_json(silent=True) or {}
    name, birth_date, username, password = payload.get("name"), payload.get("birthDate"), payload.get("username"), payload.get("password")
    if not all([name, birth_date, username, password]): return jsonify({"message": "모든 항목을 입력해 주세요."}), 400
    users = read_users()
    if any(same_text(user.get("username"), username) for user in users): return jsonify({"message": "이미 사용 중인 아이디입니다."}), 409
    users.append({"name": str(name).strip(), "birthDate": str(birth_date).strip(), "username": str(username).strip(), "password": str(password).strip(), "createdAt": str(__import__("datetime").datetime.now().isoformat(timespec="seconds"))})
    write_users(users)
    return jsonify({"message": "회원가입이 완료되었습니다."})

@app.post("/api/login")
def login():
    payload = request.get_json(silent=True) or {}
    username, password = payload.get("username"), payload.get("password")
    users = read_users()
    user = next((item for item in users if same_text(item.get("username"), username) and same_text(item.get("password"), password)), None)
    if not user: return jsonify({"message": "아이디 또는 비밀번호가 일치하지 않습니다."}), 401
    return jsonify({"message": f"{user['name']}님, 로그인되었습니다.", "profile": {"name": user["name"], "birthDate": user["birthDate"], "username": user["username"]}})

@app.post("/api/find-id")
def find_id():
    payload = request.get_json(silent=True) or {}
    name, birth_date = payload.get("name"), payload.get("birthDate")
    users = read_users()
    user = next((item for item in users if same_text(item.get("name"), name) and same_text(item.get("birthDate"), birth_date)), None)
    if not user: return jsonify({"message": "일치하는 회원정보를 찾지 못했습니다."}), 404
    return jsonify({"username": user["username"]})

@app.post("/api/find-password")
def find_password():
    payload = request.get_json(silent=True) or {}
    username, name, birth_date = payload.get("username"), payload.get("name"), payload.get("birthDate")
    users = read_users()
    user = next((item for item in users if same_text(item.get("username"), username) and same_text(item.get("name"), name) and same_text(item.get("birthDate"), birth_date)), None)
    if not user: return jsonify({"message": "입력하신 정보와 일치하는 비밀번호를 찾지 못했습니다."}), 404
    return jsonify({"password": user["password"]})

# ==========================================
# 웹소켓 (채팅, 수어, WebRTC)
# ==========================================
@socketio.on('join_room')
def on_join(data):
    room = data['room']
    name = data['name']
    join_room(room)
    emit('user_joined', {'name': name, 'sid': request.sid}, to=room, include_self=False)
    emit('receive_message', {'name': 'SYSTEM', 'msg': f'{name}님이 회의실에 입장하셨습니다.'}, to=room, include_self=True)

@socketio.on('send_message')
def on_chat_message(data): 
    emit('receive_message', data, to=data['room'], include_self=True)

@socketio.on('send_file')
def on_file_upload(data): 
    emit('receive_file', data, to=data['room'], include_self=True)

@socketio.on('webrtc_offer')
def on_webrtc_offer(data): emit('webrtc_offer', data, to=data['room'], include_self=False)

@socketio.on('webrtc_answer')
def on_webrtc_answer(data): emit('webrtc_answer', data, to=data['room'], include_self=False)

@socketio.on('webrtc_ice')
def on_webrtc_ice(data): emit('webrtc_ice', data, to=data['room'], include_self=False)

@socketio.on('sign_frame')
def on_sign_frame(data):
    room = data['room']
    name = data['name']
    image_b64 = data['image'].split(',')[1] 
    
    img_data = base64.b64decode(image_b64)
    nparr = np.frombuffer(img_data, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    # room+name은 같은 방에 같은 이름으로 접속하면 충돌할 수 있으므로,
    # 소켓 연결마다 고유한 request.sid로 세션 상태를 구분한다.
    queue_key = request.sid
    current_time = time.time()

    if queue_key not in user_finger_queues:
        user_finger_queues[queue_key] = []
        user_last_active[queue_key] = current_time
    if queue_key not in user_landmark_seqs:
        # 최대 fps(~35)로 window_seconds*3 만큼 도착해도 넉넉하도록 여유 있게 잡는다.
        user_landmark_seqs[queue_key] = deque(maxlen=128)

    detector = get_detector(queue_key)
    seq = user_landmark_seqs[queue_key]

    # 손 미검출 등으로 한동안 갱신이 없었으면, 오래된 항목은 리샘플링에 쓰이지 않도록 정리한다.
    while seq and (current_time - seq[0][0]) > WINDOW_SECONDS * 3:
        seq.popleft()

    # MediaPipe/TensorFlow는 블로킹 호출이라 eventlet 허브를 멈추게 하므로,
    # 별도 OS 스레드(tpool)에서 실행해 다른 소켓 트래픽이 계속 처리되도록 한다.
    right_hand_lmList = eventlet.tpool.execute(_detect_right_hand, detector, img)

    predicted_jamo = None

    if right_hand_lmList is not None:
        joint = np.zeros((42, 2))
        for j, lm in enumerate(right_hand_lmList):
            joint[j] = [lm.x, lm.y]

        vector, angle_label = Vector_Normalization(joint)
        seq.append((current_time, np.concatenate([vector.flatten(), angle_label.flatten()])))

        resampled = resample_time_buffer(seq, SEQ_LENGTH, WINDOW_SECONDS, current_time)
        if resampled is not None:
            input_data = np.expand_dims(resampled, axis=0)
            y_pred = eventlet.tpool.execute(_predict_gesture, input_data)
            class_id = int(np.argmax(y_pred))
            confidence = float(y_pred[class_id])

            # 신뢰도 PREDICTION_CONFIDENCE_THRESHOLD 이상만 통과 (노이즈 방어)
            if confidence > PREDICTION_CONFIDENCE_THRESHOLD:
                predicted_jamo = ACTIONS[class_id]

    queue = user_finger_queues[queue_key]
    if queue_key not in user_sentence_words:
        user_sentence_words[queue_key] = []
    sentence_words = user_sentence_words[queue_key]

    # 1. 입력 중: 한글은 자음/모음이 번갈아 나오는 구조이므로, 자음<->모음 전환이 감지될 때
    #    비로소 직전까지 표시되던 후보 음운을 큐에 확정한다. 같은 타입 내에서는 후보만 갱신된다.
    if predicted_jamo is not None:
        user_last_active[queue_key] = current_time
        pending = user_pending_jamo.get(queue_key)

        if predicted_jamo != pending:
            if pending is not None and is_vowel(predicted_jamo) != is_vowel(pending):
                queue.append(pending)
            user_pending_jamo[queue_key] = predicted_jamo

            preview_chars = queue + [predicted_jamo]
            raw_assembled = join_jamos("".join(preview_chars), ignore_err=True)
            current_word = remove_single_letters(raw_assembled) or raw_assembled
            preview_text = " ".join(sentence_words + [current_word]) if current_word else " ".join(sentence_words)

            emit('sign_progress', {
                'current_jamo': predicted_jamo,
                'text': preview_text
            }, to=request.sid)

    # 2. 입력 중단 시 (WORD_PAUSE: 단어 확정 / SENTENCE_PAUSE: 문장 전체 송출)
    else:
        time_passed = current_time - user_last_active.get(queue_key, current_time)

        if (queue or queue_key in user_pending_jamo) and time_passed >= WORD_PAUSE_SECONDS:
            # 아직 확정되지 못한 마지막 후보도 이 시점에 큐로 편입
            pending = user_pending_jamo.pop(queue_key, None)
            if pending is not None:
                queue.append(pending)

            # 음운(자모) -> 음절 조합 후, 조합되지 못한 단자음/단모음은 버리고 완성된 단어만 문장에 편입
            raw_assembled = join_jamos("".join(queue), ignore_err=True)
            finished_word = remove_single_letters(raw_assembled)
            if finished_word:
                sentence_words.append(finished_word)
            user_finger_queues[queue_key] = []

        if sentence_words and time_passed >= SENTENCE_PAUSE_SECONDS:
            # 확정된 단어들을 한 문장으로 묶어 한 번에 송출
            full_sentence = " ".join(sentence_words)
            emit('sign_result', {'name': name, 'text': full_sentence}, to=room, include_self=True)

            user_sentence_words[queue_key] = []
            emit('sign_progress', {'current_jamo': '', 'text': ''}, to=request.sid)

if __name__ == "__main__":
    ensure_workbook()
    reset_sessions()  # 서버 켤 때: 이전 실행의 잔여 세션 상태 없이 깨끗하게 시작
    atexit.register(reset_sessions)  # 서버 끌 때: MediaPipe 검출기 정리 및 세션 상태 초기화

    print("[SYSTEM] Union-web 실시간 소켓 서버를 시작합니다.")
    socketio.run(app, host="0.0.0.0", port=80, debug=False)
